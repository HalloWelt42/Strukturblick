"""Unit-Tests der XLSX-Engine: Wurzelform, Zell-Positionen, Typen, Mehrblatt, Erkennung.

Die Beispiel-Arbeitsmappen entstehen zur Laufzeit mit openpyxl in einem BytesIO -
es wird bewusst keine Binärdatei committet.
"""

from __future__ import annotations

import io
from datetime import datetime

import openpyxl  # type: ignore[import-untyped]  # kein py.typed / keine Stubs verfügbar
import pytest

from app.engines.xlsx_engine import XlsxEngine
from app.fehler import KonvertierungUnmoeglich
from app.kern.dokument import GeparstesDokument
from app.kern.erkennung import erkenne
from app.modelle.dokument import ParseOptionen, SerialisierungsOptionen
from app.modelle.gemeinsam import FormatId, Verlustaspekt
from app.registry import entdecke_module


@pytest.fixture()
def engine() -> XlsxEngine:
    return XlsxEngine()


def _einzelblatt_mappe() -> bytes:
    """Eine Arbeitsmappe mit genau einem Blatt und gemischten Zelltypen."""
    mappe = openpyxl.Workbook()
    blatt = mappe.active
    blatt.title = "Kunden"
    blatt.append(["kundennummer", "name", "umsatz", "kunde_seit", "aktiv"])
    blatt.append(["K-1001", "Erika Musterfrau", 4820.5, datetime(2020, 1, 15), True])
    blatt.append(["K-1002", "Sönke Matthiesen", 12750, datetime(2019, 6, 3), True])
    blatt.append(["K-1003", "Jürgen Möller", None, datetime(2021, 11, 22), False])
    puffer = io.BytesIO()
    mappe.save(puffer)
    return puffer.getvalue()


def _mehrblatt_mappe() -> bytes:
    """Eine Arbeitsmappe mit zwei Blättern (zweiter Blattname mit Umlaut)."""
    mappe = openpyxl.Workbook()
    kunden = mappe.active
    kunden.title = "Kunden"
    kunden.append(["id", "name"])
    kunden.append([1, "Erika"])
    kunden.append([2, "Sönke"])
    staedte = mappe.create_sheet("Städte")
    staedte.append(["stadt", "plz"])
    staedte.append(["Kiel", "24103"])
    staedte.append(["Flensburg", "24937"])
    puffer = io.BytesIO()
    mappe.save(puffer)
    return puffer.getvalue()


def test_einzelblatt_wurzel_ist_liste(engine: XlsxEngine) -> None:
    dokument = engine.parsen(_einzelblatt_mappe(), ParseOptionen())

    assert dokument.format_id == FormatId.XLSX
    assert isinstance(dokument.wurzel, list)
    assert len(dokument.wurzel) == 3
    assert dokument.wurzel[0]["kundennummer"] == "K-1001"
    assert dokument.wurzel[0]["name"] == "Erika Musterfrau"
    # Fehlender Wert (leere Zelle) -> None
    assert dokument.wurzel[2]["umsatz"] is None
    assert dokument.wurzel[2]["kundennummer"] == "K-1003"


def test_typen_bleiben_erhalten(engine: XlsxEngine) -> None:
    dokument = engine.parsen(_einzelblatt_mappe(), ParseOptionen())
    assert isinstance(dokument.wurzel, list)

    # Zahlen und Wahrheitswerte bleiben ihre nativen JSON-Typen.
    assert dokument.wurzel[0]["umsatz"] == pytest.approx(4820.5)
    assert isinstance(dokument.wurzel[0]["umsatz"], float)
    assert dokument.wurzel[1]["umsatz"] == 12750
    assert dokument.wurzel[0]["aktiv"] is True
    assert dokument.wurzel[2]["aktiv"] is False

    # datetime wird zu einem ISO-Text (mit Typpräzisions-Aspekt und Warnung).
    assert dokument.wurzel[0]["kunde_seit"] == "2020-01-15T00:00:00"
    assert Verlustaspekt.TYPPRAEZISION in dokument.genutzte_aspekte
    assert any("typpraezision" in warnung for warnung in dokument.warnungen)


def test_zell_positionen_zeile_und_spalte(engine: XlsxEngine) -> None:
    dokument = engine.parsen(_einzelblatt_mappe(), ParseOptionen())

    # Erste Datenzeile ist Tabellenzeile 2 (Kopfzeile = 1); name ist Spalte 2.
    name_pos = dokument.positionen["/0/name"]
    assert name_pos.wert.start.zeile == 2
    assert name_pos.wert.start.spalte == 2
    assert name_pos.wert.start.offset == -1  # binär: kein Zeichenoffset

    # Zweite Datenzeile (Tabellenzeile 3), umsatz ist Spalte 3.
    umsatz_pos = dokument.positionen["/1/umsatz"]
    assert umsatz_pos.wert.start.zeile == 3
    assert umsatz_pos.wert.start.spalte == 3

    # Jede Zeile trägt zusätzlich einen eigenen Pointer (Spalte 1).
    zeilen_pos = dokument.positionen["/0"]
    assert zeilen_pos.wert.start.zeile == 2
    assert zeilen_pos.wert.start.spalte == 1


def test_mehrblatt_wird_dict_mit_aspekt(engine: XlsxEngine) -> None:
    dokument = engine.parsen(_mehrblatt_mappe(), ParseOptionen())

    assert isinstance(dokument.wurzel, dict)
    assert set(dokument.wurzel.keys()) == {"Kunden", "Städte"}
    assert dokument.wurzel["Kunden"][0] == {"id": 1, "name": "Erika"}
    assert dokument.wurzel["Städte"][1] == {"stadt": "Flensburg", "plz": "24937"}

    assert Verlustaspekt.MEHRERE_TABELLEN in dokument.genutzte_aspekte
    assert any("Blätter" in warnung for warnung in dokument.warnungen)


def test_mehrblatt_positionen_mit_blattpraefix(engine: XlsxEngine) -> None:
    dokument = engine.parsen(_mehrblatt_mappe(), ParseOptionen())

    kunden_pos = dokument.positionen["/Kunden/0/name"]
    assert kunden_pos.wert.start.zeile == 2
    assert kunden_pos.wert.start.spalte == 2

    staedte_pos = dokument.positionen["/Städte/1/plz"]
    assert staedte_pos.wert.start.zeile == 3
    assert staedte_pos.wert.start.spalte == 2


def test_doppelte_spaltennamen_werden_eindeutig(engine: XlsxEngine) -> None:
    mappe = openpyxl.Workbook()
    blatt = mappe.active
    blatt.append(["name", "name", "wert"])
    blatt.append(["a", "b", "c"])
    puffer = io.BytesIO()
    mappe.save(puffer)

    dokument = engine.parsen(puffer.getvalue(), ParseOptionen())
    assert isinstance(dokument.wurzel, list)
    zeile = dokument.wurzel[0]
    assert zeile["name"] == "a"
    assert zeile["name_2"] == "b"
    assert zeile["wert"] == "c"
    assert any("Doppelter Spaltenname" in warnung for warnung in dokument.warnungen)


def test_serialisieren_ist_unmoeglich(engine: XlsxEngine) -> None:
    dokument = GeparstesDokument(format_id=FormatId.XLSX, wurzel=[{"a": 1}])

    with pytest.raises(KonvertierungUnmoeglich) as fehler_info:
        engine.serialisieren(dokument, SerialisierungsOptionen())
    assert "nur gelesen" in fehler_info.value.meldung


def test_erkennen_endung_und_magic(engine: XlsxEngine) -> None:
    roh = _einzelblatt_mappe()

    mit_endung = engine.erkennen(roh, "kunden.xlsx")
    assert mit_endung is not None
    assert mit_endung.konfidenz == pytest.approx(0.95)

    nur_inhalt = engine.erkennen(roh, None)
    assert nur_inhalt is not None
    assert nur_inhalt.konfidenz == pytest.approx(0.8)

    # Reiner Text ist kein XLSX.
    assert engine.erkennen(b"kein,zip,inhalt", None) is None


def test_erkennen_rangfolge_xlsx_zuerst() -> None:
    entdecke_module()
    kandidaten = erkenne(_einzelblatt_mappe(), "kunden.xlsx")

    assert kandidaten
    assert kandidaten[0].format_id == FormatId.XLSX
